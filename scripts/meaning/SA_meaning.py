#!/usr/bin/env python
# coding: utf-8

# Import the openpyxl library to work with Excel files
import openpyxl

# Load the source workbook
workbook = openpyxl.load_workbook('./sst_test_map_syn_t5.xlsx')
sheet = workbook.active

# Create two new workbooks and their respective sheets to store the results
res_workbook_1 = openpyxl.Workbook()
res_sheet_1 = res_workbook_1.active
res_workbook_2 = openpyxl.Workbook()
res_sheet_2 = res_workbook_2.active

# Set a threshold value to compare scores
threshold = 0.1

# Initialize variables to keep track of row indices and temporary row storage
temp_list = []
former_row = None
index_set_1 = set([])
index_set_2 = set([])
first = True

# Iterate through each row in the source sheet
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
    if first:
        first = False
        continue

    # First situation - check for rows with nonzero 'res_pos_score' or 'res_neg_score'
    if int(row[12].value) != 0 and row[0].value == former_row[0].value:
        temp_list.append(row)
    else:
        # For pairs of rows within the same 'original_text', check if their 'res_pos_score' or 'res_neg_score' difference is greater than or equal to the threshold
        for i in range(len(temp_list)-1):
            for j in range(i+1, len(temp_list)):
                if abs(temp_list[i][10].value - temp_list[j][10].value) >= threshold:
                    index_set_1.add(int(temp_list[i][12].value))
                    index_set_1.add(int(temp_list[j][12].value))
        temp_list = [row]

    # Second situation - check for rows with 'res_sentiment' equal to 1 (positive sentiment)
    if row[5].value == 1:
        # Check if the latter positive score is higher than the former one
        if row[10].value - row[2].value >= threshold:
            index_set_2.add(int(row[12].value))
    else:
        # Check if the latter negative score is higher than the former one
        if row[11].value - row[3].value >= threshold:
            index_set_2.add(int(row[12].value))

    former_row = row

# Convert the sets of row indices to lists for further processing
index_list_1 = list(index_set_1)
index_list_2 = list(index_set_2)

# Append the selected rows to the result sheets
from tqdm import tqdm
for index in tqdm(index_list_1):
    res_sheet_1.append([cell.value for cell in sheet[index+2]])

for index in tqdm(index_list_2):
    res_sheet_2.append([cell.value for cell in sheet[index+2]])

# Save the result workbooks to Excel files
res_workbook_1.save('./res_1_sentiment_meaning_t5.xlsx')
res_workbook_2.save('./res_2_sentiment_meaning_t5.xlsx')
